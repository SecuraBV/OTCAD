import openpyxl as xls
import helpers.datasets as ds
import helpers.mitre as mitre
from helpers.parseIncidents import parseIncidents
import pickle
import time
from helpers.incident import Incident
from os import path
import sys


def FilterIncidents(incidents: list[Incident]):
    ''' Filters the incidents before outputting them to either excel structure, see readme for all options.
        (Add or change filters here)
    '''
    filtered = incidents

    #Example:
    filtered = CompleteOnly(filtered)
    #filtered = SpecificAttackerType(filtered, "Targeted Attack")

    return filtered

def FilterIncidentsPerYear(incidents: list[Incident], year):
    ''' Filters the incidents for the per year excel workbook, default is cummulative up until and including the parsed year
    '''
    filtered = incidents
    #filtered = IncludeThreshold(filtered, 1)
    filtered = UptoandIncludingYear(filtered, year)

    return filtered

######################
### Usable filters ###
######################
def SpecificIndustry(incidents: list[Incident], industry: str):
    ''' Returns only attacks on a specific industry
    ''' 
    filtered = []

    if industry not in ds.industries:
        raise ValueError(f"[{industry}] is not a valid industry category, see datasets.py for the valid industry categories")

    for incident in incidents:
        if incident.industry == industry:
            filtered.append(incident)

    return filtered

def GetAttack(incidents: list[Incident], guid: str):
    ''' Gets a single attack by GUID
    '''
    filtered = 0

    for incident in incidents:
        if incident.guid == guid:
            filtered = incident

    if filtered == 0:
        raise ValueError(f"Attack with guid [{guid}] could not be found.")

    return filtered

def SpecificAttackerType(incidents: list[Incident], attackerType: str):
    ''' Only includes incidents with the provided attacker type (see helpers/datasets.py for available attacker types).
    '''
    filtered = []

    if attackerType not in ds.attackTypes and attackerType != None:
        raise ValueError(f"[{attackerType}] is not a valid attacker type, see datasets.py for the valid attacker types.")

    for incident in incidents:
        if incident.attackType == attackerType:
            filtered.append(incident)

    return filtered


def UptoandIncludingYear(incidents: list[Incident], year: int):
    ''' Only include incidents up to and including the provided year.
    '''
    filtered = []
    for incident in incidents:
        if incident.year <= year:
            filtered.append(incident)

    return filtered


def AfterAndIncludingYear(incidents: list[Incident], year: int):
    ''' Only include incidents after and including the provided year.
    '''
    filtered = []
    for incident in incidents:
        if incident.year >= year:
            filtered.append(incident)

    return filtered

def IncludeThreshold(incidents: list[Incident], threshold: int):
    ''' Include only incidents with atleast the provided amount of techniques mapped
    ''' 
    filtered = []
    for incident in incidents:
        shouldAdd = False
        count = 0
        for technique in incident.mitre:
            if incident.mitre[technique] != ["Unknown"]:
                count += 1

            if count >= threshold:
                shouldAdd = True
                break
        
        if shouldAdd:
            filtered.append(incident)
        
    return filtered


def CompleteOnly(incidents: list[Incident]):
    ''' Only include completely mapped (for all tactics atleast one mapping or not applicable) attacks
    '''
    filtered = []
    for incident in incidents:
        shouldAdd = True
        for technique in incident.mitre:
            if incident.mitre[technique] == ["Unknown"]:
                shouldAdd = False
                break
        
        if shouldAdd:
            filtered.append(incident)
        
    return filtered


def ExcludeAttackType(incidents: list[Incident], attType: str):
    ''' Exclude incidents with provided attacker type 
    '''
    filtered = []

    for incident in incidents:
        if incident.attackType != attType:
            filtered.append(incident)

    return filtered


def FilterByYear(incidents: list[Incident], year: int):
    ''' Only include incidents of provided year
    '''
    filtered = []

    for incident in incidents:
        if incident.year == year:
            filtered.append(incident)

    return filtered


def FilterByTactic(incidents: list[Incident], tactic: str):
    ''' Only include incidents with a technique mapped in provided tactic
    '''
    filtered = []

    #Check if tactic is correct
    try:
        tacticId = list(mitre.tactics.keys())[list(mitre.tactics.values()).index(tactic)]
    except:
        raise ValueError(f"[{tactic}] is not a valid tactic, see mitre.py for the valid tactics")

    for incident in incidents:
        if "Unknown" not in incident.mitre[tacticId]:
            filtered.append(incident)

    return filtered

def FilterByTechniqueAndTactic(incidents: list[Incident], tactic: str, technique: str):
    ''' Only include incidents with provided technique in provided tactic mapped
    '''
    filtered = []

    #Check if technique is correct
    if technique != "Unknown" and technique != "Not applicable":
        try:
            techniqueId = list(mitre.techniques.keys())[list(mitre.techniques.values()).index(technique)]
        except: 
            raise ValueError(f"[{technique}] is not a valid technique, see mitre.py for the valid techniques")
    else:
        techniqueId = technique

    #Check if tactic is correct
    try:
        tacticId = list(mitre.tactics.keys())[list(mitre.tactics.values()).index(tactic)]
    except:
        raise ValueError(f"[{tactic}] is not a valid tactic, see mitre.py for the valid tactics")

    #Check if technique is in tactic
    if techniqueId not in mitre.techniquesInTactic[tacticId] and techniqueId != "Unknown" and techniqueId != "Not applicable":
        raise ValueError(f"[{technique}] is not in [{tactic}], see helpers/mitre.py for the lists (the ATT&CK for ICS v8 matrix is used)")

    for incident in incidents:
        if techniqueId in incident.mitre[tacticId]:
            filtered.append(incident)

    return filtered

def ExcludeByTechniqueAndTactic(incidents: list[Incident], tactic: str, technique: str):
    ''' Exclude incidents with provided technique in provided tactic mapped
    '''
    filtered = []

    #Check if technique is correct
    if technique != "Unknown" and technique != "Not applicable":
        try:
            techniqueId = list(mitre.techniques.keys())[list(mitre.techniques.values()).index(technique)]
        except: 
            raise ValueError(f"[{technique}] is not a valid technique, see mitre.py for the valid techniques")
    else:
        techniqueId = technique

    #Check if tactic is correct
    try:
        tacticId = list(mitre.tactics.keys())[list(mitre.tactics.values()).index(tactic)]
    except:
        raise ValueError(f"[{tactic}] is not a valid tactic, see mitre.py for the valid tactics")

    #Check if technique is in tactic
    if techniqueId not in mitre.techniquesInTactic[tacticId] and techniqueId != "Unknown" and techniqueId != "Not applicable":
        raise ValueError(f"[{technique}] is not in [{tactic}], see mitre.py for the lists (the ATT&CK for ICS v8 matrix is used)")

    for incident in incidents:
        if techniqueId not in incident.mitre[tacticId]:
            filtered.append(incident)

    return filtered


###########################
# PARSING RELATED SCRIPTS #
###########################

def ExtractData(incidents: list[Incident]):
    ''' Extract the needed data from the given incidents
    '''
    #Create data holders
    industryDict = ds.getEmptyIndustryDict(0)
    attackerDict = ds.getEmptyAttackerDict(0)
    mitreDict = mitre.getEmptyTacticDict({})
    for tactic in mitreDict:
        mitreDict[tactic] = mitre.getTechniqueDict(tactic, 0)
        mitreDict[tactic]["Unknown"] = 0
        mitreDict[tactic]["Not applicable"] = 0

    #Extract data from all incidents
    for incident in incidents:
        #Add incident to the right industry and attack
        industryDict[incident.industry] += 1
        attackerDict[incident.attackType] += 1

        #Loop over all tactics and their techniques, summing the counts
        for tactic in mitreDict:
            for technique in mitre.techniquesInTactic[tactic]:
                mitreDict[tactic][technique] += incident.mitre[tactic].count(technique)

            mitreDict[tactic]["Unknown"] += incident.mitre[tactic].count("Unknown")
            mitreDict[tactic]["Not applicable"] += incident.mitre[tactic].count("Not applicable")


    return mitreDict, industryDict, attackerDict


def FillInTotalCell(sheet, labelLoc: list[int], value: int, amountIncidents: int):
    ''' Fill in excel sheet cell for the total excel workbook
    '''
    sheet.cell(column=labelLoc[0]+1, row=labelLoc[1]).value = value

    if amountIncidents == 0:
        sheet.cell(column=labelLoc[0]+2, row=labelLoc[1]).value = value
    else:
        sheet.cell(column=labelLoc[0]+2, row=labelLoc[1]).value = value / amountIncidents

    if value == 0:
        sheet.cell(column=labelLoc[0], row=labelLoc[1]).fill = xls.styles.PatternFill(fgColor="E8E8E8", fill_type = "solid")
        sheet.cell(column=labelLoc[0]+1, row=labelLoc[1]).fill = xls.styles.PatternFill(fgColor="E8E8E8", fill_type = "solid")
        sheet.cell(column=labelLoc[0]+2, row=labelLoc[1]).fill = xls.styles.PatternFill(fgColor="E8E8E8", fill_type = "solid")

def FillInYearCell(sheet, column: int, row: int, value: int):
    ''' Fill in excel sheet cell for the year excel workbook
    '''
    sheet.cell(column=column, row=row).value = value
    
    if value == 0:
        sheet.cell(column=column, row=row).fill = xls.styles.PatternFill(fgColor="E8E8E8", fill_type = "solid")

def findCellValueInSheet(sheet, value):
    ''' Returns the [column, row] of the cell with the provided value
    '''
    for column in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            cellVal = sheet.cell(column=column, row=row).value
            #Check if cell value is a string, if yes remove any new lines
            if type(cellVal) == str:
                if value == cellVal.replace("\n", ""): #This might be a problem if you run this script on Linux
                    return [column, row]
            elif value == cellVal:
                return [column, row]
    
    raise ValueError(f"[{value}] is not in sheet")


def SaveToTotalXLSX(amountIncidents: int, matrix, industries, attackers, name: str):
    ''' Save data to excel sheet with total numbers
    '''
    cellValues = {}
    #Load frame to parse data in
    workbook = xls.load_workbook(filename="excel frames/frame.xlsx")
    sheet = workbook.active

    allTactics = mitre.getReadableTactics()
    #Find all text containing cells
    for column in range(1, sheet.max_column + 1):
        currentTactic = ""
        for row in range(1, sheet.max_row + 1):
            value = sheet.cell(column=column, row=row).value

            if value in allTactics:
                currentTactic = value

            #If not none, there is text in the cell, the cells that have text are the cells we are interested in
            if value != None:
                key = value.replace("Man.", "Manufacturing").replace("\n","") #Dirty replace of Man. and removing new lines
                key += currentTactic
                cellValues[key] = [column, row]
        
    #Fill attacker types
    for attacker in attackers:
        neededCell = attacker
        if attacker == "Unknown":
            neededCell = "Attacker Unknown"

        FillInTotalCell(sheet, cellValues[neededCell], attackers[attacker], amountIncidents)

    #Fill industries
    for industry in industries:
        neededCell = industry
        if industry == "Unknown":
            neededCell = "Industry unknown"

        FillInTotalCell(sheet, cellValues[neededCell], industries[industry], amountIncidents)

    #Fill matrix
    for tactic in matrix:
        readableTactic = mitre.tactics[tactic]

        for technique in mitre.techniquesInTactic[tactic]:
            readableTechnique = mitre.techniques[technique]
            neededCell = readableTechnique + readableTactic
            FillInTotalCell(sheet, cellValues[neededCell], matrix[tactic][technique], amountIncidents)
        
        FillInTotalCell(sheet, cellValues["Unknown" + readableTactic], matrix[tactic]["Unknown"], amountIncidents)
        FillInTotalCell(sheet, cellValues["Not applicable" + readableTactic], matrix[tactic]["Not applicable"], amountIncidents)

    
    #Save workbook
    workbook.save(filename=f"{name}.xlsx")

def SaveToYearsXLSX(matrix, industries, attackers, name: str, years: list[int]):
    ''' Save data to the year excel workbook
    '''
    workbook = xls.load_workbook(filename="excel frames/year frame.xlsx")
    sheet = workbook.active

    yearCol = 2
    for year in years:
        yearMatrix = matrix[year]
        yearIndustry = industries[year]
        yearAttacker = attackers[year]

        #insert attacker types
        sheet = workbook['Attacker types']
        sheet.cell(column=yearCol, row=1).value = year
        for attacker in yearAttacker:
            attackerRow = findCellValueInSheet(sheet, attacker)[1]
            FillInYearCell(sheet, yearCol, attackerRow, yearAttacker[attacker])

        #insert industries
        sheet = workbook['Industries']
        sheet.cell(column=yearCol, row=1).value = year
        for industry in yearIndustry:
            industryRow = findCellValueInSheet(sheet, industry)[1]
            FillInYearCell(sheet, yearCol, industryRow, yearIndustry[industry])

        #insert tactics
        for tactic in yearMatrix:
            sheet = workbook[mitre.tactics[tactic]]
            sheet.cell(column=yearCol, row=1).value = year

            for technique in yearMatrix[tactic]:
                if technique == "Unknown" or technique == "Not applicable":
                    techniqueRow = findCellValueInSheet(sheet, technique)[1]
                else:
                    techniqueRow = findCellValueInSheet(sheet, mitre.techniques[technique])[1]

                FillInYearCell(sheet, yearCol, techniqueRow, yearMatrix[tactic][technique])

        yearCol += 1

    #Save workbook
    workbook.save(filename=f"{name}.xlsx")




########
# MAIN #
########
if __name__ == "__main__":

    jsonFile = "../cyberattacks.json"
    attackFolder = "../cyberattacks/"
    outputType = "total"
    outputName = "all attacks"

    if not path.isfile(jsonFile):
        print("The provided json could not be found, please confirm that you entered the file path correctly.")
        quit()

    if not path.isdir(attackFolder):
        print("The provided folder could not be found, please confirm that you entered the path correctly.")
        quit()

    if outputType != "year" and outputType != "total":
        print("please provide if you want a year overview or a total overview as output correctly.")
        quit()

    #Start
    incidents = []

    #Check if incident folder is modified since last parse
    currentModifyTime = time.ctime(path.getmtime(attackFolder))
    workingModifyTime = time.ctime(path.getmtime(jsonFile))

    if not path.isfile("cache.blob"):
        print("cache not found, creating this first by parsing the incident data.")
        parseIncidents(jsonFile, attackFolder)
        

    parsedFile = open("cache.blob", "rb")
    blobWorkingModifyTime = pickle.load(parsedFile)
    blobModifyTime = pickle.load(parsedFile)
    
    if blobModifyTime != currentModifyTime or workingModifyTime != blobWorkingModifyTime: #If true -> reparse before continueing
        print("INFO: Reparsing incident data as folder or json was updated since last parse")
        parsedFile.close()
        incidents = parseIncidents(jsonFile, attackFolder)
    else: #Continue like normal
        incidents = pickle.load(parsedFile)
        parsedFile.close()

    filteredIncidents = FilterIncidents(incidents)

    if outputType == "year":
        #Find all years that have an incident
        years = []
        for incident in incidents:
            years.append(incident.year)

        years = list(set(years))
        years.sort()

        #Loop over all years
        matrix = {}
        industries = {}
        attackers = {}
        for year in years:
            filteredIncidentsYear = FilterIncidentsPerYear(filteredIncidents, year)
            matrix[year], industries[year], attackers[year] = ExtractData(filteredIncidentsYear)
        
        SaveToYearsXLSX(matrix, industries, attackers, outputName, years) 
    else:
        matrix, industries, attackers = ExtractData(filteredIncidents)
        SaveToTotalXLSX(len(filteredIncidents), matrix, industries, attackers, outputName)
